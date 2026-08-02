"""Global pytest configuration."""

from __future__ import annotations
import pytest
from collections import defaultdict
from collections.abc import Generator, Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any, DefaultDict, cast
from _pytest.reports import TestReport
from _pytest.runner import CallInfo
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pyspark.sql import SparkSession
from _pytest.tmpdir import TempPathFactory
from pathlib import Path
from handuflow.platform.configurator import SystemConfigurator
from handuflow.platform.configurator.dataclasses.context import ConfigurationContext

# ======================================================================================
# Shared Paths
# ======================================================================================

TEST_ROOT: Path = Path(__file__).resolve().parent
REPORT_DIRECTORY: Path = TEST_ROOT / "reports"
REPORT_FILE: Path = REPORT_DIRECTORY / "pytest_report.xlsx"


@pytest.fixture(scope="session")
def test_root() -> Path:
    """
    Root directory of the test suite.
    """
    return TEST_ROOT


# ======================================================================================
# Test Result Model
# ======================================================================================


@dataclass(slots=True, frozen=True)
class TestResult:
    """Represents a single executed pytest test."""

    domain: str
    module: str
    test_file: str
    test_name: str
    status: str
    duration_seconds: float


_TEST_RESULTS: list[TestResult] = []


# ======================================================================================
# Helpers
# ======================================================================================


def _extract_domain_and_module(test_file: str) -> tuple[str, str]:
    """
    Extract the domain and module from a test path.

    Example
    -------
    tests/contract/platform/storage/test_copy.py

    Returns
    -------
    ("platform", "storage")
    """

    parts = Path(test_file).parts

    try:
        contract_index = parts.index("contract")
    except ValueError:
        return ("general", "-")

    domain = parts[contract_index + 1] if len(parts) > contract_index + 1 else "general"

    module = parts[contract_index + 2] if len(parts) > contract_index + 2 else "-"

    return (domain, module)


def _autosize_columns(worksheet: Worksheet) -> None:
    """
    Automatically size worksheet columns.
    """

    for column in worksheet.columns:
        cells = list(column)

        if not cells:
            continue

        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in cells
        )

        first_cell = cast(Cell, cells[0])
        column_letter: str = first_cell.column_letter

        worksheet.column_dimensions[column_letter].width = max_length + 4


# ======================================================================================
# Hooks
# ======================================================================================


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(
    item: pytest.Item,
    call: CallInfo[object],
) -> Generator[None, None, None]:
    """
    Capture every executed test.
    """

    outcome: Any = yield

    report = cast(TestReport, outcome.get_result())

    if report.when != "call":
        return

    domain, module = _extract_domain_and_module(item.location[0])

    _TEST_RESULTS.append(
        TestResult(
            domain=domain,
            module=module,
            test_file=Path(item.location[0]).name,
            test_name=item.name,
            status=report.outcome.upper(),
            duration_seconds=round(report.duration, 3),
        )
    )


def pytest_sessionfinish(
    session: pytest.Session,
    exitstatus: int,
) -> None:
    """
    Generate the Excel report after pytest finishes.
    """

    REPORT_DIRECTORY.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    grouped_results: DefaultDict[str, list[TestResult]] = defaultdict(list)

    for result in _TEST_RESULTS:
        grouped_results[result.domain].append(result)

    # Handle collection/import failures where no tests executed.
    if not grouped_results:
        worksheet = workbook.create_sheet("Summary")
        worksheet.append(["Status", "Message"])
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet.append(
            [
                "NO TESTS",
                "Pytest finished before executing any tests. Check the console for collection/import errors.",
            ]
        )
        _autosize_columns(worksheet)

        workbook.save(REPORT_FILE)

        print(f"\nPytest report generated: {REPORT_FILE}")
        return

    for domain in sorted(grouped_results):

        worksheet = workbook.create_sheet(title=domain)

        headers = (
            "Module",
            "Test File",
            "Test Name",
            "Status",
            "Duration (Seconds)",
        )

        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for result in grouped_results[domain]:
            worksheet.append(
                (
                    result.module,
                    result.test_file,
                    result.test_name,
                    result.status,
                    result.duration_seconds,
                )
            )

        _autosize_columns(worksheet)

    workbook.save(REPORT_FILE)

    print(f"\nPytest report generated: {REPORT_FILE}")


@pytest.fixture(scope="session")
def spark(
    tmp_path_factory: TempPathFactory,
) -> Iterator[SparkSession]:
    root: Path = tmp_path_factory.mktemp("spark")
    warehouse: Path = root / "spark-warehouse"
    metastore: Path = root / "metastore_db"
    spark = (
        SparkSession.builder.master("local[2]")
        .appName("HanduFLOW Data Quality Tests")
        .enableHiveSupport()
        .config(
            "spark.sql.warehouse.dir",
            warehouse.as_uri(),
        )
        .config(
            "javax.jdo.option.ConnectionURL",
            f"jdbc:derby:;databaseName={metastore};create=true",
        )
        .config(
            "derby.system.home",
            str(root),
        )
        .getOrCreate()
    )
    yield spark
    spark.stop()


@pytest.fixture(scope="session")
def configuration_context(
    spark: SparkSession,
) -> ConfigurationContext:

    TEST_HANDUFLOW_DIR = (
        Path(__file__).resolve().parents[0] / "test_handuflow_dir"  # tests/
    )

    print(TEST_HANDUFLOW_DIR)

    configuration = SystemConfigurator(str(TEST_HANDUFLOW_DIR), spark)
    configuration.configure()
    return configuration.get_configuration_context()
